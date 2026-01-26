# app/services/excel_parser.py
# (This file is responsible for all Excel file ingestion and parsing.)

import traceback
from flask import current_app
from app.jwt_auth import require_jwt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

# --- Service Dependencies ---
from .variables import get_latest_master_variables
from .transactions import _calculate_financial_metrics, _convert_to_json_safe


@require_jwt 
def process_excel_file(excel_file):
    """
    Orchestrates the entire process of reading, validating, and calculating data 
    from the uploaded Excel file, using master variables for key financial rates.
    """
    try:
        # Access config variables from the current Flask app context
        config = current_app.config
        
        # Helper functions for type conversion
        def safe_float(val):
            """
            Converts a value to float, treating None, empty strings, and invalid values as 0.0.
            Openpyxl returns None for empty cells instead of pandas NaN.

            When using data_only=True, Excel errors (#VALUE!, #DIV/0!, etc.) may be returned
            as strings. We log these to detect broken templates.
            """
            if val is not None and val != '':
                # Check for Excel error strings
                if isinstance(val, str) and val.startswith('#'):
                    current_app.logger.warning(f"Excel error detected in cell: {val} - Template may be broken")
                    return 0.0

                try:
                    return float(val)
                except (ValueError, TypeError):
                    current_app.logger.warning(f"Failed to convert value to float: {val} (type: {type(val).__name__})")
                    return 0.0
            return 0.0

        def safe_int(val):
            """
            Converts a value to int, treating None, empty strings, and invalid values as 0.
            """
            if val is not None and val != '':
                try:
                    return int(float(val))  # Handle "5.0" string -> 5
                except (ValueError, TypeError):
                    return 0
            return 0

        # --- NEW BLOCK: FETCH LATEST MASTER VARIABLES (Decoupling) ---
        required_master_variables = ['tipoCambio', 'costoCapital', 'tasaCartaFianza']
        latest_rates = get_latest_master_variables(required_master_variables)
        
        # Check if the necessary rates were found in the DB (CRITICAL VALIDATION)
        if (latest_rates.get('tipoCambio') is None or
            latest_rates.get('costoCapital') is None or
            latest_rates.get('tasaCartaFianza') is None):
             return {"success": False, "error": "Cannot calculate financial metrics. System rates (Tipo de Cambio, Costo Capital, or Tasa Carta Fianza) are missing. Please ensure they have been set by the Finance department."}, 400

        # --- MASTER VARIABLES SNAPSHOT: Freeze rates at upload time ---
        master_variables_snapshot = {
            'tipoCambio': latest_rates['tipoCambio'],
            'costoCapital': latest_rates['costoCapital'],
            'tasaCartaFianza': latest_rates['tasaCartaFianza'],
            'captured_at': datetime.utcnow().isoformat()
        }
        # --- END NEW BLOCK ---


        # --- PERFORMANCE OPTIMIZATION: Read Excel file with openpyxl (memory-efficient) ---
        current_app.logger.info("Reading Excel file with openpyxl (read_only mode for memory optimization)")
        excel_file.seek(0)  # Ensure we read from the start of the file stream

        workbook = None
        try:
            # Use read_only=True for memory optimization, data_only=True to get values instead of formulas
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook[config['PLANTILLA_SHEET_NAME']]  # 'PLANTILLA'

            current_app.logger.info(f"Excel sheet loaded: {worksheet.max_row} rows × {worksheet.max_column} columns")
            # ---------------------------------------------------------

            # Step 3: Read & Extract Header Data using direct cell access
            header_data = {}
            for var_name, cell_ref in config['VARIABLES_TO_EXTRACT'].items():
                # Use coordinate string directly - much cleaner than manual parsing!
                # Example: worksheet['C2'].value directly accesses cell C2
                cell_value = worksheet[cell_ref].value

                # Convert based on expected data type
                if var_name in ['MRC', 'NRC', 'plazoContrato', 'comisiones', 'companyID', 'orderID']:
                    header_data[var_name] = safe_float(cell_value)
                else:
                    header_data[var_name] = str(cell_value) if cell_value is not None else ""

            # This logic is now OVERWRITTEN by the refactor. The real commission is calculated later.
            if 'comisiones' in header_data:
                header_data['comisiones'] = 0.0

            # --- INJECT MASTER VARIABLES INTO HEADER DATA ---
            header_data['tipoCambio'] = latest_rates['tipoCambio']
            header_data['costoCapitalAnual'] = latest_rates['costoCapital']
            header_data['tasaCartaFianza'] = latest_rates['tasaCartaFianza']
            header_data['aplicaCartaFianza'] = False  # Default to NO
            header_data['master_variables_snapshot'] = master_variables_snapshot  # Frozen audit trail
            # --- END INJECTION ---

            # Extract recurring services with manual iteration (openpyxl)
            recurring_services_data = []
            services_start_row = config['RECURRING_SERVICES_START_ROW']
            services_columns = config['RECURRING_SERVICES_COLUMNS']  # {'tipo_servicio': 'J', ...}

            empty_row_count = 0
            MAX_EMPTY_ROWS = 5  # Stop after 5 consecutive empty rows

            # Iterate from start row to max_row
            for row_idx in range(services_start_row + 1, worksheet.max_row + 1):  # +1 for 1-based indexing
                row_data = {}
                is_empty_row = True

                # Extract each column value
                for field_name, col_letter in services_columns.items():
                    col_idx = column_index_from_string(col_letter)
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value

                    # Track if row has any non-empty cells
                    # IMPORTANT: Strip whitespace - a cell with " " should be treated as empty
                    if cell_value is not None and str(cell_value).strip() != '':
                        is_empty_row = False

                    row_data[field_name] = cell_value

                # Skip completely empty rows (equivalent to dropna(how='all'))
                if is_empty_row:
                    empty_row_count += 1
                    if empty_row_count >= MAX_EMPTY_ROWS:
                        break  # Stop reading after 5 consecutive empty rows
                else:
                    empty_row_count = 0  # Reset counter
                    recurring_services_data.append(row_data)

            current_app.logger.info(f"SUCCESS: Read {len(recurring_services_data)} recurring service records")

            # Extract fixed costs with manual iteration (openpyxl)
            fixed_costs_data = []
            fixed_costs_start_row = config['FIXED_COSTS_START_ROW']
            fixed_costs_columns = config['FIXED_COSTS_COLUMNS']  # {'categoria': 'A', ...}

            current_app.logger.info(f"--- DEBUG: Fixed Costs Extraction ---")
            current_app.logger.info(f"Starting from row: {fixed_costs_start_row + 1}")
            current_app.logger.info(f"Expected columns: {len(fixed_costs_columns)}")

            empty_row_count = 0
            MAX_EMPTY_ROWS = 5  # Stop after 5 consecutive empty rows

            # Iterate from start row to max_row
            for row_idx in range(fixed_costs_start_row + 1, worksheet.max_row + 1):
                row_data = {}
                is_empty_row = True

                # Extract each column value
                for field_name, col_letter in fixed_costs_columns.items():
                    col_idx = column_index_from_string(col_letter)
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value

                    # Track if row has any non-empty cells
                    # IMPORTANT: Strip whitespace - a cell with " " should be treated as empty
                    if cell_value is not None and str(cell_value).strip() != '':
                        is_empty_row = False

                    row_data[field_name] = cell_value

                # Skip completely empty rows
                if is_empty_row:
                    empty_row_count += 1
                    if empty_row_count >= MAX_EMPTY_ROWS:
                        break  # Stop reading after 5 consecutive empty rows
                else:
                    empty_row_count = 0  # Reset counter
                    fixed_costs_data.append(row_data)

            current_app.logger.info(f"SUCCESS: Read {len(fixed_costs_data)} fixed cost records")
            current_app.logger.info(f"--- END DEBUG ---\n")

            # Calculate totals for preview
            for item in fixed_costs_data:
                # Rename to _original pattern
                item['costoUnitario_original'] = safe_float(item.get('costoUnitario', 0))
                item['costoUnitario_currency'] = 'USD'

                # Calculate total for preview (in original currency)
                cantidad = item.get('cantidad')
                costo_original = item.get('costoUnitario_original')
                if cantidad is not None and costo_original is not None:
                    item['total'] = cantidad * costo_original

                item['periodo_inicio'] = safe_float(item.get('periodo_inicio', 0))
                item['duracion_meses'] = safe_float(item.get('duracion_meses', 1))

            for item in recurring_services_data:
                q = safe_float(item.get('Q', 0))
                p_original = safe_float(item.get('P', 0))
                cu1_original = safe_float(item.get('CU1', 0))
                cu2_original = safe_float(item.get('CU2', 0))

                # Rename to _original pattern
                item['P_original'] = p_original
                item['P_currency'] = 'PEN'
                item['CU1_original'] = cu1_original
                item['CU2_original'] = cu2_original
                item['CU_currency'] = 'USD'

                # Calculate preview values in original currency
                item['ingreso'] = q * p_original
                item['egreso'] = (cu1_original + cu2_original) * q

            # <-- MODIFIED: This is the total in *original* currency, not PEN
            calculated_costoInstalacion = sum(
                item.get('total', 0) for item in fixed_costs_data if item.get('total') is not None)

            # Step 4: Validate Inputs
            client_name = header_data.get('clientName')
            mrc_value = header_data.get('MRC')
            if not client_name or client_name == '' or mrc_value is None or mrc_value == '':
                return {"success": False, "error": "Required field 'Client Name' or 'MRC' is missing from the Excel file."}

            # Rename to _original pattern for transaction
            header_data['MRC_original'] = header_data.get('MRC')
            header_data['MRC_currency'] = 'PEN'
            header_data['NRC_original'] = header_data.get('NRC')
            header_data['NRC_currency'] = 'PEN'

            # Consolidate all extracted data
            full_extracted_data = {**header_data, 'recurring_services': recurring_services_data,
                                   'fixed_costs': fixed_costs_data, 'costoInstalacion': calculated_costoInstalacion}

            # Step 5: Calculate Metrics
            # This function now calculates *all* metrics, including the *real* commission.
            # It needs the GIGALAN/Unidad fields, but they are not in the Excel file.
            # They will be None, so commission will correctly calculate as 0.0 for now.
            # This is the *correct* initial state.
            # <-- This function now handles all PEN conversions internally
            financial_metrics = _calculate_financial_metrics(full_extracted_data)

            # Step 6: Assemble the Final Response
            # <-- MODIFIED: 'costoInstalacion' is now the PEN-based value from financial_metrics
            transaction_summary = {
                **header_data,
                **financial_metrics,
                "costoInstalacion": financial_metrics.get('costoInstalacion'), # This is now PEN
                "submissionDate": None,
                "ApprovalStatus": "PENDING"
            }

            final_data_package = {"transactions": transaction_summary, "fixed_costs": fixed_costs_data,
                                  "recurring_services": recurring_services_data}

            clean_data = _convert_to_json_safe(final_data_package)

            return {"success": True, "data": clean_data}

        finally:
            # Always close the workbook to free resources
            if workbook:
                workbook.close()
                current_app.logger.info("Workbook closed successfully")

    except Exception as e:
        import traceback
        print("--- ERROR DURING EXCEL PROCESSING ---")
        print(traceback.format_exc())
        print("--- END ERROR ---")
        return {"success": False, "error": f"An unexpected error occurred: {str(e)}"}